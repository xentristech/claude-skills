#!/usr/bin/env python3
"""
Genera el worksheet de reconocimiento SEO en formato Excel (.xlsx), con las 4 hojas
de la plantilla: Reconocimiento, Contenidos, Recomendaciones y Enlaces.

USO:
    python generar_worksheet_excel.py salida.xlsx --sitio "https://ejemplo.com"

Luego se rellena con los hallazgos (a mano o programáticamente). Requiere openpyxl:
    pip install openpyxl
"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEAD = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="2F5496")
SECTION = PatternFill("solid", fgColor="D9E1F2")
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = FILL


def build(sitio: str) -> Workbook:
    wb = Workbook()

    # --- Hoja 1: Reconocimiento ---
    ws = wb.active
    ws.title = "Reconocimiento"
    ws.append(["Sitio web", "Tráfico orgánico actual", "Total términos de búsqueda", "Autoridad del dominio"])
    _style_header(ws, 1, 4)
    ws.append([sitio, "", "", ""])
    ws.append([])

    cols = ["Ítem", "Estado", "Detalles del hallazgo", "Documentación de apoyo"]

    def section(title, items):
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=1).fill = SECTION
        ws.append(cols)
        _style_header(ws, ws.max_row, 4)
        for it in items:
            ws.append([it, "", "", ""])
        ws.append([])

    section("Reconocimiento técnico", [
        "El sitio web cuenta con un archivo Robots.txt",
        "Las reglas del Robots.txt NO impiden rastrear contenido importante",
        "El sitio web tiene un Sitemap en formato XML",
        "El sitio web está indexado en Google",
        "URLs importantes son indexables (sin noindex indebido)",
        "URLs importantes con profundidad de navegación ≤ 3",
    ])
    section("Factores críticos para el algoritmo y las personas", [
        "El sitio web tiene una versión optimizada para móviles",
        "Las páginas principales NO tienen errores de usabilidad en móvil",
        "El sitio web carga por la versión segura (HTTPS)",
    ])
    section("Factores de carga del sitio frente a Google", [
        "Los elementos del sitio web cargan para Google",
        "Dependencia de JavaScript para renderizar el contenido",
        "La velocidad de carga del sitio web es inferior a 4 segundos",
    ])
    section("Reconocimiento estratégico", [
        "Secciones posicionadas más importantes del sitio web",
        "Qué tipos de contenido tiene el sitio web",
        "Qué tipos de intención traen más tráfico al sitio web",
        "Quiénes son los principales competidores",
        "Qué ganancias prontas se pueden identificar de los competidores",
    ])

    # --- Hoja 2: Contenidos ---
    ws2 = wb.create_sheet("Contenidos")
    campos = ["URL", "Término de búsqueda principal", "Términos de búsqueda secundarios",
              "Formato de contenido", "Título SEO", "Meta descripción",
              "Título principal del contenido (H1)", "Recomendaciones de estructura de contenido"]
    ws2.append(["Campo", "Valor propuesto"])
    _style_header(ws2, 1, 2)
    for f in campos:
        ws2.append([f, ""])

    # --- Hoja 3: Recomendaciones ---
    ws3 = wb.create_sheet("Recomendaciones")
    ws3.append(["Prioridad", "Acción recomendada", "Impacto", "Esfuerzo", "Responsable", "Apoyo"])
    _style_header(ws3, 1, 6)

    # --- Hoja 4: Enlaces ---
    ws4 = wb.create_sheet("Enlaces")
    ws4.append(["No.", "Keyword", "Prospecto (URL)", "DA", "Medio de contacto",
                "Fecha de contacto", "Respuesta"])
    _style_header(ws4, 1, 7)

    # Anchos y wrap razonables
    for sheet in wb.worksheets:
        for col in sheet.columns:
            letter = col[0].column_letter
            sheet.column_dimensions[letter].width = 32
            for cell in col:
                cell.alignment = WRAP
    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("salida", help="Ruta del .xlsx de salida")
    ap.add_argument("--sitio", default="", help="URL del sitio analizado")
    args = ap.parse_args()
    wb = build(args.sitio)
    wb.save(args.salida)
    print(f"Worksheet generado: {args.salida}")


if __name__ == "__main__":
    main()
